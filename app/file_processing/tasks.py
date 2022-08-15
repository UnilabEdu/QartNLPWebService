import json
import os
from itertools import islice

import textract
from openpyxl import load_workbook
from ftfy import fix_encoding
from werkzeug.utils import secure_filename

from app.database import db
from app.extensions import celery
from app.file_processing.nlp import *
from app.models.file import File, Pages, Sentences, Words, Statistics
from app.settings import basedir

from flask_login import current_user


@celery.task()
def process_excel_file(filename, user_id, sheet_name, start_row, word_column, lemma_column):
    from app import create_app
    with create_app().app_context():
        path = os.path.join(Config.UPLOAD_FOLDER, user_id, filename)
        excel_file = load_workbook(path)
        sheet = excel_file[sheet_name]

        for index in range(start_row, sheet.max_row + 1):
            word = sheet[word_column + str(index)].value
            lemma = lemmatize([word])[0][1]
            sheet[lemma_column + str(index)] = lemma
        
        name, extension = filename.split('.')
        path = os.path.join(Config.UPLOAD_FOLDER, user_id, name + "-COMPLETED." + extension)
        excel_file.save(path)


@celery.task()
def process_file(id, user, filename, processes, extension):
    from app import create_app
    with create_app().app_context():

        file_path = os.path.join(Config.UPLOAD_FOLDER, str(user), filename)
        available_extensions = ['docx', 'html', 'pdf']
        if extension in available_extensions:
            filename = filename.replace('.', '_') + '_converted.txt'
            file_object = File.query.get(id)
            file_object.file_name = filename

            if extension == 'html':
                current_file = open(file_path, 'r', encoding='utf-8')
                plain_text = current_file.read()
                current_file.close()
            else:
                plain_text = textract.process(file_path, input_encoding='utf-8', output_encoding='utf-8').decode('utf-8')

            converted_txt_path = os.path.join(basedir, 'uploads', str(user), filename)

            with open(converted_txt_path, "w", encoding='utf-8') as text_file:
                text_file.write(plain_text)
                text_file.close()
            converted_txt_path = os.path.join(basedir, 'uploads', str(user), filename)

            with open(converted_txt_path, "w", encoding='utf-8') as text_file:
                text_file.write(plain_text)
                text_file.close()

            file_path = converted_txt_path
        # modifying/cleaning and replacing text in the uploaded file
        current_file = open(file_path, 'r', encoding='utf-8')
        current_text = fix_encoding(current_file.read())
        current_file.close()
        if 'remove_html' in processes or extension == 'html':
            current_text = remove_html_tags(current_text)
        if 'clean_whitespaces' in processes:
            current_text = remove_trailing_spaces(current_text)
        if 'clean_special_characters' in processes:
            current_text = remove_special_characters(current_text)

        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(current_text)
        # generating frequency distribution and lemma tags
        if "freq_dist" in processes:

            freq_file = open(file_path, 'r', encoding='utf-8')
            freq_text = fix_encoding(freq_file.read())
            freq_file.close()

            result_json = frequency_distribution(freq_text)
            freq_data = json.dumps(result_json, ensure_ascii=False, indent=1)
            newtitle = f"{filename}-freq_dist.json"
            freq_filepath = os.path.join(Config.UPLOAD_FOLDER, str(user), newtitle)
            with open(freq_filepath, "w", encoding='utf-8') as fp:
                fp.write(freq_data)

        if "lemat" in processes:
            with open(file_path, "r", encoding='utf-8') as file:
                page_start = 0
                while True:
                    text = fix_encoding(''.join(islice(file, 75)))
                    if not text:
                        break
                    page_end = page_start + len(text)
                    page_db = Pages(id, page_start, page_end)
                    page_db.flush()

                    list_of_sentences = split_sentences(text)
                    sentence_start = 0
                    for sentences in list_of_sentences:
                        sentence_end = sentence_start + len(sentences)
                        sentence_db = Sentences(page_db.id, sentence_start, sentence_end)
                        sentence_db.flush()

                        lemmatized_words = lemmatize(tokenize(remove_punctuation(sentences)))
                        for word in lemmatized_words:
                            words_db = Words(sentence_db.id, word[3], word[4], word[0], word[1], word[2])
                            words_db.flush()

                        sentence_start = sentence_end + 1

                    page_start = page_end
                db.session.commit()

            file = File.file_by_id(id)
            statistics = Statistics(id, file.get_word_count(), file.get_unique_word_count(), file.get_sentence_count(),
                                    None, None)
            statistics.save()
            db.session.commit()

        File.file_by_id(id).status[0].completed = True
        db.session.commit()
